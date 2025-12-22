# ========================================================================
# MISSION DATA EXCEL EXPORT MODULE
# ========================================================================
"""
Excel export functionality for complete mission analysis data.

Export structure:
    - Complete Mission: Concatenated time-series X(t) for all phases
    - Individual phases: Climb, Cruise, Descent trajectories
    - Mission Summary: Aggregate statistics (Σm_fuel, Σt, Σs by phase)
    - Optimization History: Convergence data (fuel/range optimizers)
    - CG Analysis: x_CG(t) evolution and tank distribution m_i(t)

Output format: Multi-sheet Excel workbook with auto-sized columns and headers.
"""

from __future__ import annotations
import numpy as np
import pandas as pd
from pathlib import Path
from typing import Optional, Dict, Any, List

# Aircraft geometry and atmospheric models
from aircraft_config import S_REF_M2
from atmosphere import a_from_altitude, isa_properties

# Directory management
from visualization_config import get_or_create_run_directory

# Excel formatting utilities
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment

# CG system interface
from cg_x_calculation import _get_fuel_system, TANK_CG_POSITIONS, TANK_NAMES

# ========================================================================
# SECTION 1: MAIN EXPORT INTERFACE
# ========================================================================

def export_mission_to_excel(
    climb_result,
    cruise_result,
    descent_result,
    initial_mass_kg: float,
    output_dir: Optional[str] = None,
    fuel_optimization_history=None,
    range_optimization_history=None,
    climb_info: Optional[Dict[str, Any]] = None,
    descent_info: Optional[Dict[str, Any]] = None
) -> str:
    """
    Export complete mission trajectory data to multi-sheet Excel workbook.
    
    Workbook structure:
        Sheet 1: Complete Mission - Concatenated X(t) for all phases
        Sheet 2: Climb Phase - X_climb(t), performance metrics
        Sheet 3: Cruise Phase - X_cruise(t), performance metrics
        Sheet 4: Descent Phase - X_descent(t), performance metrics
        Sheet 5: Mission Summary - Aggregate statistics by phase
        Sheet 6: Fuel Optimization - Convergence history (optional)
        Sheet 7: Range Optimization - Iteration history (optional)
        Sheet 8: CG Analysis - x_CG(t), tank distribution m_i(t) (optional)
    
    Time alignment: Cumulative values (t, s, m_fuel) continuous across phases.
    
    Parameters:
        climb_result: MinFuelSchedule - climb trajectory
        cruise_result: CruiseResults - cruise trajectory
        descent_result: DescentResults - descent trajectory
        initial_mass_kg: m_0 [kg] - initial aircraft mass
        output_dir: str - output directory (default: workspace/simulation data)
        fuel_optimization_history: ConvergenceHistory - fuel optimizer data (optional)
        range_optimization_history: List - range optimizer data (optional)
        climb_info: dict - climb metadata (optional)
        descent_info: dict - descent metadata (optional)
    
    Returns:
        str: path to saved Excel file
    """
    
    # ────────────────────────────────────────────────────────────────────
    # Output Directory Resolution
    # ────────────────────────────────────────────────────────────────────
    if output_dir is None:
        workspace_root = Path(__file__).parent
        output_dir = workspace_root / "simulation data"
        output_dir.mkdir(parents=True, exist_ok=True)
    else:
        output_dir = Path(output_dir)
        output_dir.mkdir(parents=True, exist_ok=True)
    
    # ────────────────────────────────────────────────────────────────────
    # Phase Data Preparation
    # ────────────────────────────────────────────────────────────────────
    climb_data = _prepare_climb_data(climb_result, initial_mass_kg)
    cruise_data = _prepare_cruise_data(cruise_result)
    descent_data = _prepare_descent_data(descent_result)
    
    # ────────────────────────────────────────────────────────────────────
    # Cumulative Variable Alignment Across Phases
    # ────────────────────────────────────────────────────────────────────
    # Ensure continuity: t_cruise = t_climb,f + Δt_cruise
    # Climb terminal values: t_climb,f, s_climb,f, m_fuel,climb
    climb_end_time = float(climb_data['Cumulative Time (s)'].iloc[-1]) if len(climb_data) > 0 and 'Cumulative Time (s)' in climb_data.columns else 0.0
    climb_end_distance = float(climb_data['Cumulative Distance (km)'].iloc[-1]) if len(climb_data) > 0 and 'Cumulative Distance (km)' in climb_data.columns else 0.0
    climb_end_fuel = float(climb_data['Cumulative Fuel (kg)'].iloc[-1]) if len(climb_data) > 0 and 'Cumulative Fuel (kg)' in climb_data.columns else 0.0
    
    # Cruise phase offset: t_cruise = t_climb,f + Δt_cruise
    if len(cruise_data) > 0:
        cruise_data = cruise_data.copy()
        if 'Cumulative Time (s)' in cruise_data.columns:
            cruise_data['Cumulative Time (s)'] = cruise_data['Cumulative Time (s)'] + climb_end_time
        if 'Cumulative Distance (km)' in cruise_data.columns:
            cruise_data['Cumulative Distance (km)'] = cruise_data['Cumulative Distance (km)'] + climb_end_distance
        if 'Cumulative Fuel (kg)' in cruise_data.columns:
            cruise_data['Cumulative Fuel (kg)'] = cruise_data['Cumulative Fuel (kg)'] + climb_end_fuel
    
    # Cruise terminal values: t_cruise,f, s_cruise,f, m_fuel,cruise
    cruise_end_time = float(cruise_data['Cumulative Time (s)'].iloc[-1]) if len(cruise_data) > 0 and 'Cumulative Time (s)' in cruise_data.columns else climb_end_time
    cruise_end_distance = float(cruise_data['Cumulative Distance (km)'].iloc[-1]) if len(cruise_data) > 0 and 'Cumulative Distance (km)' in cruise_data.columns else climb_end_distance
    cruise_end_fuel = float(cruise_data['Cumulative Fuel (kg)'].iloc[-1]) if len(cruise_data) > 0 and 'Cumulative Fuel (kg)' in cruise_data.columns else climb_end_fuel
    
    # Descent phase offset: t_descent = t_cruise,f + Δt_descent
    if len(descent_data) > 0:
        descent_data = descent_data.copy()
        if 'Cumulative Time (s)' in descent_data.columns:
            descent_data['Cumulative Time (s)'] = descent_data['Cumulative Time (s)'] + cruise_end_time
        if 'Cumulative Distance (km)' in descent_data.columns:
            descent_data['Cumulative Distance (km)'] = descent_data['Cumulative Distance (km)'] + cruise_end_distance
        if 'Cumulative Fuel (kg)' in descent_data.columns:
            descent_data['Cumulative Fuel (kg)'] = descent_data['Cumulative Fuel (kg)'] + cruise_end_fuel
    
    # ────────────────────────────────────────────────────────────────────
    # DataFrame Construction
    # ────────────────────────────────────────────────────────────────────
    # Concatenate phases: X_mission(t) = [X_climb, X_cruise, X_descent]
    all_data = pd.concat([climb_data, cruise_data, descent_data], ignore_index=True)
    
    # Ensure chronological ordering
    if 'Cumulative Time (s)' in all_data.columns:
        all_data = all_data.sort_values('Cumulative Time (s)').reset_index(drop=True)
    
    # Mission summary statistics
    summary_data = _create_mission_summary(
        climb_data, cruise_data, descent_data, 
        climb_result, cruise_result, descent_result,
        initial_mass_kg, climb_info, descent_info
    )
    
    # Optional: Optimization convergence histories
    fuel_opt_data = None
    if fuel_optimization_history and len(fuel_optimization_history.iterations) > 0:
        fuel_opt_data = _prepare_fuel_optimization_data(fuel_optimization_history)
    
    range_opt_data = None
    if range_optimization_history and len(range_optimization_history) > 0:
        range_opt_data = _prepare_range_optimization_data(range_optimization_history)
    
    # Optional: CG evolution data
    cg_data = _prepare_cg_data()
    
    # ────────────────────────────────────────────────────────────────────
    # Excel Workbook Generation
    # ────────────────────────────────────────────────────────────────────
    from datetime import datetime
    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    excel_filename = f"mission_data_{timestamp}.xlsx"
    excel_path = Path(output_dir) / excel_filename
    
    with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
        # Sheet 1: Complete mission
        all_data.to_excel(writer, sheet_name='Complete Mission', index=False)
        
        # Sheets 2-4: Individual phases
        climb_data.to_excel(writer, sheet_name='Climb Phase', index=False)
        cruise_data.to_excel(writer, sheet_name='Cruise Phase', index=False)
        descent_data.to_excel(writer, sheet_name='Descent Phase', index=False)
        
        # Sheet 5: Mission summary
        summary_data.to_excel(writer, sheet_name='Mission Summary', index=False)
        
        # Sheets 6-7: Optimization histories (conditional)
        if fuel_opt_data is not None:
            fuel_opt_data.to_excel(writer, sheet_name='Fuel Optimization', index=False)
        
        if range_opt_data is not None:
            range_opt_data.to_excel(writer, sheet_name='Range Optimization', index=False)
        
        # Sheet 8: CG analysis (conditional)
        if cg_data is not None:
            cg_data.to_excel(writer, sheet_name='CG Analysis', index=False)
        
        # Apply formatting
        _auto_size_excel_columns(writer)
    
    print(f"\n[EXCEL EXPORT] Mission data exported to: {excel_path}")
    print(f"  Total rows (Complete Mission): {len(all_data)}")
    print(f"  Climb rows: {len(climb_data)}")
    print(f"  Cruise rows: {len(cruise_data)}")
    print(f"  Descent rows: {len(descent_data)}")
    print(f"  Sheets created: {'Complete Mission, Climb Phase, Cruise Phase, Descent Phase, Mission Summary'}")
    if fuel_opt_data is not None:
        print(f"    + Fuel Optimization ({len(fuel_opt_data)} iterations)")
    if range_opt_data is not None:
        print(f"    + Range Optimization ({len(range_opt_data)} iterations)")
    if cg_data is not None:
        print(f"    + CG Analysis ({len(cg_data)} history points)")
    
    return str(excel_path)


# ========================================================================
# SECTION 2: SUMMARY STATISTICS GENERATION
# ========================================================================

def _create_mission_summary(
    climb_data: pd.DataFrame,
    cruise_data: pd.DataFrame,
    descent_data: pd.DataFrame,
    climb_result,
    cruise_result,
    descent_result,
    initial_mass_kg: float,
    climb_info: Optional[Dict[str, Any]] = None,
    descent_info: Optional[Dict[str, Any]] = None
) -> pd.DataFrame:
    """
    Generate mission summary with aggregate statistics by phase.
    
    Summary structure:
        - Mass: m_0, m_f, Δm_total
        - By phase: Δs, Δm_fuel, Δt for climb/cruise/descent
        - Total mission: Σs, Σm_fuel, Σt
        - Optimization metadata (if available)
    
    Returns:
        DataFrame: {Parameter, Value, Unit} rows
    """
    
    # Calculate phase statistics - use correct column names
    climb_fuel = float(climb_data['Cumulative Fuel (kg)'].iloc[-1]) if len(climb_data) > 0 and 'Cumulative Fuel (kg)' in climb_data.columns else 0.0
    cruise_fuel = float(cruise_data['Cumulative Fuel (kg)'].iloc[-1]) - climb_fuel if len(cruise_data) > 0 and 'Cumulative Fuel (kg)' in cruise_data.columns else 0.0
    descent_fuel = float(descent_data['Cumulative Fuel (kg)'].iloc[-1]) - (climb_fuel + cruise_fuel) if len(descent_data) > 0 and 'Cumulative Fuel (kg)' in descent_data.columns else 0.0
    total_fuel = climb_fuel + cruise_fuel + descent_fuel
    
    climb_time = float(climb_data['Cumulative Time (s)'].iloc[-1]) if len(climb_data) > 0 and 'Cumulative Time (s)' in climb_data.columns else 0.0
    cruise_time = float(cruise_data['Cumulative Time (s)'].iloc[-1]) - climb_time if len(cruise_data) > 0 and 'Cumulative Time (s)' in cruise_data.columns else 0.0
    descent_time = float(descent_data['Cumulative Time (s)'].iloc[-1]) - (climb_time + cruise_time) if len(descent_data) > 0 and 'Cumulative Time (s)' in descent_data.columns else 0.0
    total_time = climb_time + cruise_time + descent_time
    
    climb_distance = float(climb_data['Cumulative Distance (km)'].iloc[-1]) if len(climb_data) > 0 and 'Cumulative Distance (km)' in climb_data.columns else 0.0
    cruise_distance = float(cruise_data['Cumulative Distance (km)'].iloc[-1]) - climb_distance if len(cruise_data) > 0 and 'Cumulative Distance (km)' in cruise_data.columns else 0.0
    descent_distance = float(descent_data['Cumulative Distance (km)'].iloc[-1]) - (climb_distance + cruise_distance) if len(descent_data) > 0 and 'Cumulative Distance (km)' in descent_data.columns else 0.0
    total_distance = climb_distance + cruise_distance + descent_distance
    
    final_mass = float(descent_data['Mass (kg)'].iloc[-1]) if len(descent_data) > 0 and 'Mass (kg)' in descent_data.columns else initial_mass_kg - total_fuel
    
    summary_rows = [
        {'Parameter': 'Initial Mass', 'Value': f'{initial_mass_kg:.1f}', 'Unit': 'kg'},
        {'Parameter': 'Final Mass', 'Value': f'{final_mass:.1f}', 'Unit': 'kg'},
        {'Parameter': 'Mass Reduction (Fuel)', 'Value': f'{initial_mass_kg - final_mass:.1f}', 'Unit': 'kg'},
        {'Parameter': '', 'Value': '', 'Unit': ''},
        {'Parameter': 'CLIMB PHASE', 'Value': '', 'Unit': ''},
        {'Parameter': '  Distance', 'Value': f'{climb_distance:.1f}', 'Unit': 'km'},
        {'Parameter': '  Fuel Consumed', 'Value': f'{climb_fuel:.1f}', 'Unit': 'kg'},
        {'Parameter': '  Time', 'Value': f'{climb_time/3600:.2f}', 'Unit': 'hours'},
        {'Parameter': '  Final Altitude', 'Value': f'{climb_data["Altitude (m)"].iloc[-1]:.0f}', 'Unit': 'm'} if 'Altitude (m)' in climb_data.columns else {'Parameter': '  Final Altitude', 'Value': 'N/A', 'Unit': 'm'},
        {'Parameter': '  Final Mach', 'Value': f'{climb_data["Mach Number"].iloc[-1]:.3f}', 'Unit': ''} if 'Mach Number' in climb_data.columns else {'Parameter': '  Final Mach', 'Value': 'N/A', 'Unit': ''},
        {'Parameter': '', 'Value': '', 'Unit': ''},
        {'Parameter': 'CRUISE PHASE', 'Value': '', 'Unit': ''},
        {'Parameter': '  Distance', 'Value': f'{cruise_distance:.0f}', 'Unit': 'km'},
        {'Parameter': '  Fuel Consumed', 'Value': f'{cruise_fuel:.1f}', 'Unit': 'kg'},
        {'Parameter': '  Time', 'Value': f'{cruise_time/3600:.2f}', 'Unit': 'hours'},
        {'Parameter': '  Average Fuel Flow', 'Value': f'{cruise_fuel/(cruise_time/3600):.0f}' if cruise_time > 0 else '0', 'Unit': 'kg/h'},
        {'Parameter': '  Final Altitude', 'Value': f'{cruise_data["Altitude (m)"].iloc[-1]:.0f}', 'Unit': 'm'} if 'Altitude (m)' in cruise_data.columns else {'Parameter': '  Final Altitude', 'Value': 'N/A', 'Unit': 'm'},
        {'Parameter': '  Final Mach', 'Value': f'{cruise_data["Mach Number"].iloc[-1]:.3f}', 'Unit': ''} if 'Mach Number' in cruise_data.columns else {'Parameter': '  Final Mach', 'Value': 'N/A', 'Unit': ''},
        {'Parameter': '', 'Value': '', 'Unit': ''},
        {'Parameter': 'DESCENT PHASE', 'Value': '', 'Unit': ''},
        {'Parameter': '  Distance', 'Value': f'{descent_distance:.1f}', 'Unit': 'km'},
        {'Parameter': '  Fuel Consumed', 'Value': f'{descent_fuel:.1f}', 'Unit': 'kg'},
        {'Parameter': '  Time', 'Value': f'{descent_time/3600:.2f}', 'Unit': 'hours'},
        {'Parameter': '  Final Altitude', 'Value': f'{descent_data["Altitude (m)"].iloc[-1]:.0f}', 'Unit': 'm'} if 'Altitude (m)' in descent_data.columns else {'Parameter': '  Final Altitude', 'Value': 'N/A', 'Unit': 'm'},
        {'Parameter': '  Final Mach', 'Value': f'{descent_data["Mach Number"].iloc[-1]:.3f}', 'Unit': ''} if 'Mach Number' in descent_data.columns else {'Parameter': '  Final Mach', 'Value': 'N/A', 'Unit': ''},
        {'Parameter': '', 'Value': '', 'Unit': ''},
        {'Parameter': 'TOTAL MISSION', 'Value': '', 'Unit': ''},
        {'Parameter': '  Total Distance', 'Value': f'{total_distance:.1f}', 'Unit': 'km'},
        {'Parameter': '  Total Fuel Consumed', 'Value': f'{total_fuel:.1f}', 'Unit': 'kg'},
        {'Parameter': '  Total Time', 'Value': f'{total_time/3600:.2f}', 'Unit': 'hours'},
        {'Parameter': '  Fuel as % of Initial Mass', 'Value': f'{total_fuel/initial_mass_kg*100:.1f}', 'Unit': '%'},
    ]
    
    # Add optimization info if available
    if climb_info:
        summary_rows.append({'Parameter': '', 'Value': '', 'Unit': ''})
        summary_rows.append({'Parameter': 'CLIMB OPTIMIZATION INFO', 'Value': '', 'Unit': ''})
        if 'total_fuel_kg' in climb_info:
            summary_rows.append({'Parameter': '  Total Fuel (from info)', 'Value': f'{climb_info["total_fuel_kg"]:.1f}', 'Unit': 'kg'})
        if 'total_time_s' in climb_info:
            summary_rows.append({'Parameter': '  Total Time (from info)', 'Value': f'{climb_info["total_time_s"]/3600:.2f}', 'Unit': 'hours'})
        if 'final_mach' in climb_info:
            summary_rows.append({'Parameter': '  Final Mach (from info)', 'Value': f'{climb_info["final_mach"]:.3f}', 'Unit': ''})
        if 'final_altitude' in climb_info:
            summary_rows.append({'Parameter': '  Final Altitude (from info)', 'Value': f'{climb_info["final_altitude"]:.0f}', 'Unit': 'm'})
        if 'path_length' in climb_info:
            summary_rows.append({'Parameter': '  Path Length', 'Value': f'{climb_info["path_length"]}', 'Unit': 'points'})
    
    if descent_info:
        summary_rows.append({'Parameter': '', 'Value': '', 'Unit': ''})
        summary_rows.append({'Parameter': 'DESCENT OPTIMIZATION INFO', 'Value': '', 'Unit': ''})
        if 'total_fuel_kg' in descent_info:
            summary_rows.append({'Parameter': '  Total Fuel (from info)', 'Value': f'{descent_info["total_fuel_kg"]:.1f}', 'Unit': 'kg'})
        if 'total_time_s' in descent_info:
            summary_rows.append({'Parameter': '  Total Time (from info)', 'Value': f'{descent_info["total_time_s"]/3600:.2f}', 'Unit': 'hours'})
        if 'total_time_min' in descent_info:
            summary_rows.append({'Parameter': '  Total Time (min)', 'Value': f'{descent_info["total_time_min"]:.1f}', 'Unit': 'min'})
        if 'final_mach' in descent_info:
            summary_rows.append({'Parameter': '  Final Mach (from info)', 'Value': f'{descent_info["final_mach"]:.3f}', 'Unit': ''})
        if 'final_altitude' in descent_info:
            summary_rows.append({'Parameter': '  Final Altitude (from info)', 'Value': f'{descent_info["final_altitude"]:.0f}', 'Unit': 'm'})
        if 'target_mach' in descent_info:
            summary_rows.append({'Parameter': '  Target Mach', 'Value': f'{descent_info["target_mach"]:.3f}', 'Unit': ''})
        if 'target_altitude' in descent_info:
            summary_rows.append({'Parameter': '  Target Altitude', 'Value': f'{descent_info["target_altitude"]:.0f}', 'Unit': 'm'})
        if 'mach_deviation' in descent_info:
            summary_rows.append({'Parameter': '  Mach Deviation', 'Value': f'{descent_info["mach_deviation"]:.4f}', 'Unit': ''})
        if 'path_length' in descent_info:
            summary_rows.append({'Parameter': '  Path Length', 'Value': f'{descent_info["path_length"]}', 'Unit': 'points'})
        if 'avg_descent_rate_mps' in descent_info:
            summary_rows.append({'Parameter': '  Avg Descent Rate (m/s)', 'Value': f'{descent_info["avg_descent_rate_mps"]:.2f}', 'Unit': 'm/s'})
        if 'avg_descent_rate_mpm' in descent_info:
            summary_rows.append({'Parameter': '  Avg Descent Rate (m/min)', 'Value': f'{descent_info["avg_descent_rate_mpm"]:.1f}', 'Unit': 'm/min'})
    
    return pd.DataFrame(summary_rows)


# ========================================================================
# SECTION 3: OPTIMIZATION HISTORY PREPARATION
# ========================================================================

def _prepare_fuel_optimization_data(convergence_history) -> pd.DataFrame:
    """
    Extract fuel optimization convergence history.
    
    Data: Iteration sequence showing m_fuel evolution via bisection.
    Columns: iteration, m_fuel,initial, Δm_fuel, deficit, bounds
    
    Returns:
        DataFrame: convergence history by iteration
    """
    rows = []
    for i, result in enumerate(convergence_history.iterations):
        bounds = convergence_history.fuel_bounds_history[i] if i < len(convergence_history.fuel_bounds_history) else (0.0, 0.0)
        
        row = {
            'Iteration': result.iteration,
            'Initial Fuel (kg)': result.initial_fuel_kg,
            'Initial Mass (kg)': result.initial_mass_kg,
            'Fuel Consumed (kg)': result.fuel_consumed_kg,
            'Fuel Deficit (kg)': result.fuel_deficit_kg,
            'Lower Bound (kg)': bounds[0],
            'Upper Bound (kg)': bounds[1],
            'Bound Range (kg)': bounds[1] - bounds[0],
            'Climb Fuel (kg)': result.climb_fuel_kg,
            'Cruise Fuel (kg)': result.cruise_fuel_kg,
            'Descent Fuel (kg)': result.descent_fuel_kg,
            'Climb Time (s)': result.climb_time_s,
            'Cruise Time (s)': result.cruise_time_s,
            'Descent Time (s)': result.descent_time_s,
            'Total Time (s)': result.total_time_s,
            'Final Mass (kg)': result.final_mass_kg,  # Renamed for physics accuracy
        }
        rows.append(row)
    
    return pd.DataFrame(rows)


def _prepare_range_optimization_data(iteration_history: List) -> pd.DataFrame:
    """
    Extract range optimization iteration history.
    
    Data: Iteration sequence showing s_cruise adjustment to match s_target.
    Columns: iteration, s_cruise, s_total, error, convergence flag
    
    Returns:
        DataFrame: iteration history with distance matching
    """
    rows = []
    for record in iteration_history:
        row = {
            'Iteration': record.iteration,
            'Cruise Distance (km)': record.cruise_distance_km,
            'Total Distance (km)': record.total_distance_km,
            'Distance Error (km)': record.distance_error_km,
            'Converged': record.converged,
            'Cruise Final Mass (kg)': record.cruise_final_mass_kg,
            'Descent Initial Mass (kg)': record.descent_initial_mass_kg,
            'Descent Final Mass (kg)': record.descent_final_mass_kg,
        }
        rows.append(row)
    
    return pd.DataFrame(rows)


# ========================================================================
# SECTION 4: EXCEL FORMATTING UTILITIES
# ========================================================================

def _auto_size_excel_columns(writer):
    """
    Apply column auto-sizing and header formatting to all sheets.
    
    Operations:
        - Header row: Bold font, centered alignment
        - Column width: Auto-sized based on content (max 50 chars)
        - Sampling: First 100 rows for performance
    """
    workbook = writer.book
    
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        
        # Format header row
        header_font = Font(bold=True, size=11)
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        for cell in worksheet[1]:  # First row is header
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Auto-size columns
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            # Check header length
            if column[0].value:
                max_length = len(str(column[0].value))
            
            # Check data length (sample first 100 rows for performance)
            for cell in column[1:101]:  # Sample first 100 data rows
                try:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    pass
            
            # Set column width with some padding
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            worksheet.column_dimensions[column_letter].width = adjusted_width


# ========================================================================
# SECTION 5: PHASE DATA PREPARATION
# ========================================================================

def _prepare_climb_data(climb_result, initial_mass_kg: float) -> pd.DataFrame:
    """
    Extract and compute climb phase trajectory data.
    
    Computations:
        - Cumulative time: t = Σ Δt
        - Cumulative distance: s = Σ V·Δt
        - True airspeed: V = M·a(h)
        - Atmospheric properties: T(h), ρ(h), p(h)
        - Aerodynamic coefficients: CL, CD, L/D
        - Performance metrics: Ps, ṁ, J
    
    Returns:
        DataFrame: complete climb trajectory with derived quantities
    """
    n_points = len(climb_result.alt_m)
    
    # Cumulative time: t = Σ Δt
    time_s = np.cumsum(np.nan_to_num(climb_result.dt_s, nan=0.0, posinf=0.0, neginf=0.0))
    if len(time_s) > 0:
        time_s = time_s - time_s[0]  # Ensure t_0 = 0
    
    # Cumulative distance: s = Σ V·Δt where V = M·a(h)
    distance_km = np.zeros(n_points)
    cumulative_distance_km = 0.0
    for i in range(n_points):
        if i < len(climb_result.mach) and i < len(climb_result.alt_m) and i < len(climb_result.dt_s):
            a = a_from_altitude(float(climb_result.alt_m[i]))
            V_tas = float(climb_result.mach[i]) * a
            segment_distance_km = V_tas * float(climb_result.dt_s[i]) / 1000.0
            cumulative_distance_km += segment_distance_km
            distance_km[i] = cumulative_distance_km
    
    # True airspeed: V = M·a(h)
    true_airspeed_mps = np.zeros(n_points)
    for i in range(n_points):
        if i < len(climb_result.mach) and i < len(climb_result.alt_m):
            a = a_from_altitude(float(climb_result.alt_m[i]))
            true_airspeed_mps[i] = float(climb_result.mach[i]) * a
    
    # Atmospheric properties via ISA: T(h), ρ(h), p(h)
    atmospheric_data = [isa_properties(float(h)) for h in climb_result.alt_m]
    temperature_K = np.array([T for T, _, _ in atmospheric_data])
    density_kgpm3 = np.array([rho for _, _, rho in atmospheric_data])
    
    # Mass: m(t) from dynamic tracking or m = m_0 - Σm_fuel
    weight_kg = np.asarray(climb_result.mass_kg, float) if hasattr(climb_result, 'mass_kg') else initial_mass_kg - np.asarray(climb_result.cumFuel_kg, float)
    
    # Fuel flow: ṁ [kg/s]
    fuel_flow_kgps = np.asarray(climb_result.mdot_kgps, float) if hasattr(climb_result, 'mdot_kgps') else np.zeros(n_points)
    
    # Climb rate: ḣ = Δh/Δt
    climb_rate_mps = np.zeros(n_points)
    for i in range(1, n_points):
        if i < len(climb_result.alt_m) and i < len(time_s):
            dt = time_s[i] - time_s[i-1] if time_s[i] > time_s[i-1] else 0.001
            dh = float(climb_result.alt_m[i]) - float(climb_result.alt_m[i-1])
            climb_rate_mps[i] = dh / dt if dt > 0 else 0.0
    
    # ────────────────────────────────────────────────────────────────────
    # Derived Aerodynamic Quantities
    # ────────────────────────────────────────────────────────────────────
    # Dynamic pressure: q = 0.5·ρ·V²
    dynamic_pressure_Pa = 0.5 * density_kgpm3 * true_airspeed_mps**2
    
    # Lift coefficient: CL = L/(q·S) where L ≈ mg
    lift_N = weight_kg * 9.80665
    lift_coefficient = lift_N / (dynamic_pressure_Pa * S_REF_M2) if dynamic_pressure_Pa[0] > 0 else np.zeros(n_points)
    
    # Drag coefficient: CD = D/(q·S)
    drag_coefficient = np.asarray(climb_result.D_N, float) / (dynamic_pressure_Pa * S_REF_M2) if dynamic_pressure_Pa[0] > 0 else np.zeros(n_points)
    
    # Lift-to-drag ratio: L/D
    ld_ratio = lift_N / np.asarray(climb_result.D_N, float) if len(climb_result.D_N) > 0 else np.zeros(n_points)
    
    # Get T_per_engine_N directly from climb_result if available
    thrust_per_engine_N = np.asarray(climb_result.T_per_engine_N, float) if hasattr(climb_result, 'T_per_engine_N') else np.asarray(climb_result.thrust_total_N, float) / 2.0
    
    # Pressure (from atmospheric properties - already computed)
    pressure_Pa = np.array([p for _, p, _ in [isa_properties(float(h)) for h in climb_result.alt_m]])
    
    # Speed of sound (from atmospheric properties)
    speed_of_sound_mps = np.array([a_from_altitude(float(h)) for h in climb_result.alt_m])
    
    # J cost (directly from climb_result)
    j_cost = np.asarray(climb_result.J_kg_per_m, float) if hasattr(climb_result, 'J_kg_per_m') else np.zeros(n_points)
    
    # Thrust limited flag (directly from climb_result)
    thrust_limited = np.asarray(climb_result.thrust_limited, int) if hasattr(climb_result, 'thrust_limited') else np.zeros(n_points, dtype=int)
    
    # Fuel estimate (directly from climb_result)
    fuel_est_kg = np.full(n_points, float(climb_result.fuel_est_kg)) if hasattr(climb_result, 'fuel_est_kg') else np.zeros(n_points)
    
    # ────────────────────────────────────────────────────────────────────
    # Additional Computed Metrics for Consistency
    # ────────────────────────────────────────────────────────────────────
    # Calculate average metrics
    total_fuel = float(climb_result.cumFuel_kg[-1]) if len(climb_result.cumFuel_kg) > 0 else 0.0
    total_time = float(time_s[-1]) if len(time_s) > 0 else 0.0
    avg_thrust = float(np.mean(climb_result.thrust_total_N)) if len(climb_result.thrust_total_N) > 0 else 0.0
    avg_fuel_flow = float(np.mean(fuel_flow_kgps)) if len(fuel_flow_kgps) > 0 else 0.0
    
    # TSFC: TSFC = ṁ/T [kg/(N·s)]
    thrust_array = np.asarray(climb_result.thrust_total_N, float)
    tsfc_kg_per_N_s = np.zeros(n_points)
    for i in range(n_points):
        if thrust_array[i] > 0 and i < len(fuel_flow_kgps) and fuel_flow_kgps[i] > 0:
            tsfc_kg_per_N_s[i] = fuel_flow_kgps[i] / thrust_array[i]
    
    # Fuel efficiency: η = Δm_fuel/Δs [kg/km]
    fuel_efficiency_kg_per_km = np.zeros(n_points)
    for i in range(1, n_points):
        if i < len(distance_km) and distance_km[i] > distance_km[i-1]:
            dist_diff = distance_km[i] - distance_km[i-1]
            fuel_diff = float(climb_result.cumFuel_kg[i]) - float(climb_result.cumFuel_kg[i-1])
            fuel_efficiency_kg_per_km[i] = fuel_diff / dist_diff if dist_diff > 0 else 0.0
    
    # Distance increment (non-cumulative)
    distance_increment_km = np.concatenate([[0.0], np.diff(distance_km)])
    
    # Mission parameters
    initial_altitude = float(climb_result.alt_m[0]) if len(climb_result.alt_m) > 0 else 0.0
    initial_mach = float(climb_result.mach[0]) if len(climb_result.mach) > 0 else 0.0
    initial_mass = float(weight_kg[0]) if len(weight_kg) > 0 else initial_mass_kg
    target_altitude = float(climb_result.alt_m[-1]) if len(climb_result.alt_m) > 0 else 0.0
    target_mach = float(climb_result.mach[-1]) if len(climb_result.mach) > 0 else 0.0
    final_mass = float(weight_kg[-1]) if len(weight_kg) > 0 else initial_mass_kg - total_fuel
    target_distance_km = float(distance_km[-1]) if len(distance_km) > 0 else 0.0  # Actual distance covered during climb
    
    # Create DataFrame with all available data from climb_result
    # Organized into logical groups: Phase ID -> Time -> Distance -> Position -> Mass -> Forces -> Fuel -> Atmospheric -> Performance -> Engine -> Efficiency -> Mission Parameters
    data = {
        # Phase identification
        'Phase': ['Climb'] * n_points,
        'Strategy Name': ['Climb'] * n_points,
        
        # Time data
        'Cumulative Time (s)': time_s,
        'Time Step (s)': np.asarray(climb_result.dt_s, float),
        'Total Time (s)': np.full(n_points, total_time),
        
        # Distance data
        'Cumulative Distance (km)': distance_km,
        'Distance (km)': distance_increment_km,
        
        # Position/State
        'Altitude (m)': np.asarray(climb_result.alt_m, float),
        'Mach Number': np.asarray(climb_result.mach, float),
        'True Airspeed (m/s)': true_airspeed_mps,
        'Speed of Sound (m/s)': speed_of_sound_mps,
        
        # Mass
        'Mass (kg)': np.asarray(climb_result.mass_kg, float) if hasattr(climb_result, 'mass_kg') else weight_kg,
        
        # Forces
        'Thrust Total (N)': np.asarray(climb_result.thrust_total_N, float),
        'Thrust Per Engine (N)': thrust_per_engine_N,
        'Average Thrust (N)': np.full(n_points, avg_thrust),
        'Drag (N)': np.asarray(climb_result.D_N, float),
        'Lever Position': np.asarray(climb_result.lever, float),
        
        # Fuel data
        'Fuel Flow (kg/s)': fuel_flow_kgps,
        'Fuel Flow (kg/h)': fuel_flow_kgps * 3600.0,
        'Mass Flow Rate (kg/s)': fuel_flow_kgps,  # Same as fuel flow
        'Average Fuel Flow (kg/s)': np.full(n_points, avg_fuel_flow),
        'Cumulative Fuel (kg)': np.asarray(climb_result.cumFuel_kg, float),
        'Total Fuel Consumed (kg)': np.full(n_points, total_fuel),
        'Fuel Increment (kg)': np.asarray(climb_result.dFuel_kg, float),
        'Fuel Estimate (kg)': fuel_est_kg,
        
        # Atmospheric properties
        'Temperature (K)': temperature_K,
        'Pressure (Pa)': pressure_Pa,
        'Density (kg/m³)': density_kgpm3,
        
        # Performance metrics
        'Specific Excess Power (m/s)': np.asarray(climb_result.Ps_mps, float),
        'Climb Rate (m/s)': climb_rate_mps,
        'Descent Rate (m/s)': np.zeros(n_points),  # Not applicable for climb
        'Average Descent Rate (m/s)': np.zeros(n_points),  # Not applicable for climb
        'Average Descent Rate (m/min)': np.zeros(n_points),  # Not applicable for climb
        'J Cost (kg/m)': j_cost,
        'Thrust Limited': thrust_limited,
        
        # Aerodynamic coefficients
        'Dynamic Pressure (Pa)': dynamic_pressure_Pa,
        'Lift Coefficient': lift_coefficient,
        'Drag Coefficient': drag_coefficient,
        'L/D Ratio': ld_ratio,
        
        # Engine performance
        'TSFC (kg/(N·s))': tsfc_kg_per_N_s,
        'TSFC (g/(kN·s))': tsfc_kg_per_N_s * 1e6,  # Convert to g/(kN·s)
        
        # Force and power metrics
        'Net Force (N)': np.asarray(climb_result.thrust_total_N, float) - np.asarray(climb_result.D_N, float),
        'Thrust Power (W)': np.asarray(climb_result.thrust_total_N, float) * true_airspeed_mps,
        'Drag Power (W)': np.asarray(climb_result.D_N, float) * true_airspeed_mps,
        'Excess Power (W)': (np.asarray(climb_result.thrust_total_N, float) - np.asarray(climb_result.D_N, float)) * true_airspeed_mps,
        'Thrust Power (kW)': (np.asarray(climb_result.thrust_total_N, float) * true_airspeed_mps) / 1000.0,
        'Drag Power (kW)': (np.asarray(climb_result.D_N, float) * true_airspeed_mps) / 1000.0,
        
        # Efficiency metrics
        'Fuel Efficiency (kg/km)': fuel_efficiency_kg_per_km,
        
        # Mission parameters
        'Initial Altitude (m)': np.full(n_points, initial_altitude),
        'Initial Mach': np.full(n_points, initial_mach),
        'Initial Mass (kg)': np.full(n_points, initial_mass),
        'Target Altitude (m)': np.full(n_points, target_altitude),
        'Target Mach': np.full(n_points, target_mach),
        'Target Distance (km)': np.full(n_points, target_distance_km),  # Actual distance covered during climb
        'Final Mass (kg)': np.full(n_points, final_mass),
    }
    
    return pd.DataFrame(data)


def _prepare_cruise_data(cruise_result) -> pd.DataFrame:
    """
    Extract cruise phase trajectory data with field normalization.
    
    Computations:
        - Distance: s(t) from cruise_result
        - Atmospheric: T(h), ρ(h), p(h) from ISA
        - Aerodynamic coefficients: CL, CD, L/D, q (dynamic pressure)
        - Engine: TSFC = ṁ/T
        - Power metrics: P_thrust, P_drag, P_excess
    
    Returns:
        DataFrame: complete cruise trajectory with derived quantities
    """
    n_points = len(cruise_result.time_s)
    
    # Extract state variables
    cumulative_distance_km = np.asarray(cruise_result.distance_km, float)  # s(t) [km]
    
    # Climb/descent rates from specific excess power
    # Even in "level" flight, Ps shows slight climb/descent tendencies
    Ps_array = np.asarray(cruise_result.Ps_mps, float)
    climb_rate_mps = np.where(Ps_array > 0, Ps_array, 0.0)      # Ps > 0: slight climb tendency
    descent_rate_mps = np.where(Ps_array < 0, np.abs(Ps_array), 0.0)  # Ps < 0: slight descent tendency
    
    # Extract performance arrays from cruise_result
    weight_array = np.asarray(cruise_result.mass_kg, float)            # m(t) [kg]
    drag_array = np.asarray(cruise_result.D_N, float)               # D(t) [N]
    tas_array = np.asarray(cruise_result.true_airspeed_mps, float)     # V(t) [m/s]
    density_array = np.asarray(cruise_result.density_kgpm3, float)     # ρ(t) [kg/m³]
    thrust_array = np.asarray(cruise_result.thrust_total_N, float)     # T(t) [N]
    
    # Atmospheric properties via ISA
    pressure_Pa = np.array([p for _, p, _ in [isa_properties(float(h)) for h in cruise_result.altitude_m]])    # p(h) [Pa]
    speed_of_sound_mps = np.array([a_from_altitude(float(h)) for h in cruise_result.altitude_m])              # a(h) [m/s]
    
    # ────────────────────────────────────────────────────────────────────
    # Derived Aerodynamic Quantities
    # ────────────────────────────────────────────────────────────────────
    # Dynamic pressure: q = 0.5·ρ·V²
    dynamic_pressure_Pa = 0.5 * density_array * tas_array**2
    
    # Lift coefficient: CL = L/(q·S) where L = mg (level flight)
    lift_N = weight_array * 9.80665
    lift_coefficient = lift_N / (dynamic_pressure_Pa * S_REF_M2) if dynamic_pressure_Pa[0] > 0 else np.zeros(n_points)
    
    # Drag coefficient: CD = D/(q·S)
    drag_coefficient = drag_array / (dynamic_pressure_Pa * S_REF_M2) if dynamic_pressure_Pa[0] > 0 else np.zeros(n_points)
    
    # Lift-to-drag ratio: L/D
    ld_ratio = lift_N / drag_array if drag_array[0] > 0 else np.zeros(n_points)
    
    # TSFC: TSFC = ṁ/T [kg/(N·s)]
    tsfc_kg_per_N_s = np.zeros(n_points)
    for i in range(n_points):
        if thrust_array[i] > 0 and i < len(cruise_result.mdot_kgps):
            tsfc_kg_per_N_s[i] = cruise_result.mdot_kgps[i] / thrust_array[i] if thrust_array[i] > 0 else 0.0
    
    # Power metrics: P = F·V
    net_force_N = thrust_array - drag_array      # F_net = T - D [N]
    thrust_power_W = thrust_array * tas_array    # P_thrust = T·V [W]
    drag_power_W = drag_array * tas_array        # P_drag = D·V [W]
    excess_power_W = net_force_N * tas_array     # P_excess = (T-D)·V [W]
    
    # Fuel efficiency: η = Δm_fuel/Δs [kg/km]
    fuel_efficiency_kg_per_km = np.zeros(n_points)
    for i in range(1, n_points):
        if i < len(cumulative_distance_km) and cumulative_distance_km[i] > cumulative_distance_km[i-1]:
            dist_diff = cumulative_distance_km[i] - cumulative_distance_km[i-1]
            fuel_diff = cruise_result.fuel_consumed_kg[i] - cruise_result.fuel_consumed_kg[i-1] if i < len(cruise_result.fuel_consumed_kg) else 0.0
            fuel_efficiency_kg_per_km[i] = fuel_diff / dist_diff if dist_diff > 0 else 0.0
    
    # Get initial state data if available
    initial_altitude = cruise_result.initial_state.altitude_m if hasattr(cruise_result, 'initial_state') else cruise_result.altitude_m[0]
    initial_mach = cruise_result.initial_state.mach if hasattr(cruise_result, 'initial_state') else cruise_result.mach_number[0]
    initial_mass = cruise_result.initial_state.mass_kg if hasattr(cruise_result, 'initial_state') else weight_array[0]
    target_distance = cruise_result.target_distance_km if hasattr(cruise_result, 'target_distance_km') else cumulative_distance_km[-1]
    time_step = cruise_result.time_step_s if hasattr(cruise_result, 'time_step_s') else 0.0
    
    # Calculate additional metrics for consistency
    total_fuel = float(cruise_result.fuel_consumed_kg[-1]) if len(cruise_result.fuel_consumed_kg) > 0 else 0.0
    total_time = float(cruise_result.time_s[-1]) if len(cruise_result.time_s) > 0 else 0.0
    avg_thrust = float(cruise_result.average_thrust_N) if hasattr(cruise_result, 'average_thrust_N') else float(np.mean(thrust_array))
    avg_fuel_flow = float(cruise_result.average_fuel_flow_kgps) if hasattr(cruise_result, 'average_fuel_flow_kgps') else float(np.mean(cruise_result.mdot_kgps))
    
    # Thrust per engine (2 engines assumed)
    thrust_per_engine_N = thrust_array / 2.0
    
    # Target altitude and mach (same as cruise altitude for level flight)
    target_altitude = float(cruise_result.altitude_m[-1]) if len(cruise_result.altitude_m) > 0 else initial_altitude
    target_mach = float(cruise_result.mach_number[-1]) if len(cruise_result.mach_number) > 0 else initial_mach
    final_mass = float(weight_array[-1]) if len(weight_array) > 0 else initial_mass - total_fuel
    
    # Distance increment (non-cumulative distance per segment)
    distance_increment_km = np.concatenate([[0.0], np.diff(cumulative_distance_km)])
    
    # Thrust limited flag: Check if lever is at maximum (≈ 1.0)
    # Using same tolerance as climb module
    lever_array = np.asarray(cruise_result.lever_position, float)
    thrust_limited = np.isclose(lever_array, 1.0, atol=0.01).astype(int)
    
    # J Cost for cruise: fuel cost per meter of horizontal distance
    # J_cruise = ṁ/V = Fuel Efficiency [kg/km] / 1000 = [kg/m]
    j_cost_cruise = fuel_efficiency_kg_per_km / 1000.0  # Convert kg/km to kg/m
    
    # Create DataFrame with all available data from cruise_result
    # Organized into logical groups: Phase ID -> Time -> Distance -> Position -> Mass -> Forces -> Fuel -> Atmospheric -> Performance -> Engine -> Efficiency -> Mission Parameters
    data = {
        # Phase identification
        'Phase': ['Cruise'] * n_points,
        'Strategy Name': ['Cruise'] * n_points,
        
        # Time data
        'Cumulative Time (s)': np.asarray(cruise_result.time_s, float),
        'Time Step (s)': np.full(n_points, time_step),
        'Total Time (s)': np.full(n_points, total_time),
        
        # Distance data
        'Cumulative Distance (km)': cumulative_distance_km,
        'Distance (km)': distance_increment_km,
        
        # Position/State
        'Altitude (m)': np.asarray(cruise_result.altitude_m, float),
        'Mach Number': np.asarray(cruise_result.mach_number, float),
        'True Airspeed (m/s)': tas_array,
        'Speed of Sound (m/s)': speed_of_sound_mps,
        
        # Mass
        'Mass (kg)': weight_array,
        
        # Forces
        'Thrust Total (N)': thrust_array,
        'Thrust Per Engine (N)': thrust_per_engine_N,
        'Average Thrust (N)': np.full(n_points, avg_thrust),
        'Drag (N)': drag_array,
        'Lever Position': np.asarray(cruise_result.lever_position, float),
        
        # Fuel data
        'Fuel Flow (kg/s)': np.asarray(cruise_result.mdot_kgps, float),
        'Fuel Flow (kg/h)': np.asarray(cruise_result.mdot_kgps, float) * 3600.0,
        'Mass Flow Rate (kg/s)': np.asarray(cruise_result.mdot_kgps, float),  # Same as fuel flow
        'Average Fuel Flow (kg/s)': np.full(n_points, avg_fuel_flow),
        'Cumulative Fuel (kg)': np.asarray(cruise_result.fuel_consumed_kg, float),
        'Total Fuel Consumed (kg)': np.full(n_points, total_fuel),
        'Fuel Increment (kg)': np.concatenate([[0.0], np.diff(np.asarray(cruise_result.fuel_consumed_kg, float))]),
        'Fuel Estimate (kg)': np.full(n_points, total_fuel),  # Total fuel consumed for cruise phase
        
        # Atmospheric properties
        'Temperature (K)': np.asarray(cruise_result.temperature_K, float),
        'Pressure (Pa)': pressure_Pa,
        'Density (kg/m³)': density_array,
        
        # Performance metrics
        'Specific Excess Power (m/s)': np.asarray(cruise_result.Ps_mps, float),
        'Climb Rate (m/s)': climb_rate_mps,
        'Descent Rate (m/s)': descent_rate_mps,
        'Average Descent Rate (m/s)': np.full(n_points, float(np.mean(descent_rate_mps[descent_rate_mps > 0])) if np.any(descent_rate_mps > 0) else 0.0),
        'Average Descent Rate (m/min)': np.full(n_points, float(np.mean(descent_rate_mps[descent_rate_mps > 0]) * 60.0) if np.any(descent_rate_mps > 0) else 0.0),
        'J Cost (kg/m)': j_cost_cruise,  # Fuel cost per meter of horizontal distance
        'Thrust Limited': thrust_limited,  # Calculated from lever position
        
        # Aerodynamic coefficients
        'Dynamic Pressure (Pa)': dynamic_pressure_Pa,
        'Lift Coefficient': lift_coefficient,
        'Drag Coefficient': drag_coefficient,
        'L/D Ratio': ld_ratio,
        
        # Engine performance
        'TSFC (kg/(N·s))': tsfc_kg_per_N_s,
        'TSFC (g/(kN·s))': tsfc_kg_per_N_s * 1e6,  # Convert to g/(kN·s)
        
        # Force and power metrics
        'Net Force (N)': net_force_N,
        'Thrust Power (W)': thrust_power_W,
        'Drag Power (W)': drag_power_W,
        'Excess Power (W)': excess_power_W,
        'Thrust Power (kW)': thrust_power_W / 1000.0,
        'Drag Power (kW)': drag_power_W / 1000.0,
        
        # Efficiency metrics
        'Fuel Efficiency (kg/km)': fuel_efficiency_kg_per_km,
        
        # Mission parameters
        'Initial Altitude (m)': np.full(n_points, initial_altitude),
        'Initial Mach': np.full(n_points, initial_mach),
        'Initial Mass (kg)': np.full(n_points, initial_mass),
        'Target Altitude (m)': np.full(n_points, target_altitude),
        'Target Mach': np.full(n_points, target_mach),
        'Target Distance (km)': np.full(n_points, target_distance),
        'Final Mass (kg)': np.full(n_points, final_mass),
    }
    
    return pd.DataFrame(data)


def _prepare_descent_data(descent_result) -> pd.DataFrame:
    """
    Extract descent phase trajectory data with field normalization.
    
    Computations:
        - Cumulative time: t = Σ Δt (if not provided)
        - Cumulative distance: s = Σ V·Δt
        - True airspeed: V = M·a(h)
        - Atmospheric properties: T(h), ρ(h), p(h)
        - Aerodynamic coefficients: CL, CD, L/D
        - Performance metrics: Ps, ṁ, descent rate
    
    Returns:
        DataFrame: complete descent trajectory with derived quantities
    """
    n_points = len(descent_result.alt_m)
    
    # Cumulative time: t = Σ Δt or directly from time_s
    if len(descent_result.time_s) > 0:
        time_s = np.asarray(descent_result.time_s, float)
        if len(time_s) > 0 and time_s[0] != 0:
            time_s = time_s - time_s[0]  # Ensure t_0 = 0
    else:
        time_s = np.cumsum(np.nan_to_num(descent_result.dt_s, nan=0.0, posinf=0.0, neginf=0.0))
        if len(time_s) > 0:
            time_s = time_s - time_s[0]
    
    # Cumulative distance: s = Σ V·Δt where V = M·a(h)
    distance_km = np.zeros(n_points)
    cumulative_distance_km = 0.0
    for i in range(n_points):
        if i < len(descent_result.mach) and i < len(descent_result.alt_m) and i < len(descent_result.dt_s):
            a = a_from_altitude(float(descent_result.alt_m[i]))
            V_tas = float(descent_result.mach[i]) * a
            segment_distance_km = V_tas * float(descent_result.dt_s[i]) / 1000.0
            cumulative_distance_km += segment_distance_km
            distance_km[i] = cumulative_distance_km
    
    # Climb rate: Zero during descent (aircraft is descending, not climbing)
    climb_rate_mps = np.zeros(n_points)
    
    # Extract performance arrays from descent_result
    weight_array = np.asarray(descent_result.mass_kg, float)            # m(t) [kg]
    drag_array = np.asarray(descent_result.D_N, float)               # D(t) [N]
    tas_array = np.asarray(descent_result.true_airspeed_mps, float)     # V(t) [m/s]
    density_array = np.asarray(descent_result.density_kgpm3, float)     # ρ(t) [kg/m³]
    thrust_array = np.asarray(descent_result.thrust_total_N, float)     # T(t) [N]
    
    # Atmospheric properties via ISA
    pressure_Pa = np.array([p for _, p, _ in [isa_properties(float(h)) for h in descent_result.alt_m]])     # p(h) [Pa]
    speed_of_sound_mps = np.array([a_from_altitude(float(h)) for h in descent_result.alt_m])                # a(h) [m/s]
    
    # ────────────────────────────────────────────────────────────────────
    # Derived Aerodynamic Quantities
    # ────────────────────────────────────────────────────────────────────
    # Dynamic pressure: q = 0.5·ρ·V²
    dynamic_pressure_Pa = 0.5 * density_array * tas_array**2
    
    # Lift coefficient: CL = L/(q·S) where L ≈ mg
    lift_N = weight_array * 9.80665
    lift_coefficient = lift_N / (dynamic_pressure_Pa * S_REF_M2) if dynamic_pressure_Pa[0] > 0 else np.zeros(n_points)
    
    # Drag coefficient: CD = D/(q·S)
    drag_coefficient = drag_array / (dynamic_pressure_Pa * S_REF_M2) if dynamic_pressure_Pa[0] > 0 else np.zeros(n_points)
    
    # Lift-to-drag ratio: L/D
    ld_ratio = lift_N / drag_array if drag_array[0] > 0 else np.zeros(n_points)
    
    # TSFC: TSFC = ṁ/T [kg/(N·s)]
    tsfc_kg_per_N_s = np.zeros(n_points)
    for i in range(n_points):
        if thrust_array[i] > 0 and i < len(descent_result.mdot_kgps):
            tsfc_kg_per_N_s[i] = descent_result.mdot_kgps[i] / thrust_array[i] if thrust_array[i] > 0 else 0.0
    
    # Power metrics: P = F·V
    net_force_N = thrust_array - drag_array      # F_net = T - D [N]
    thrust_power_W = thrust_array * tas_array    # P_thrust = T·V [W]
    drag_power_W = drag_array * tas_array        # P_drag = D·V [W]
    excess_power_W = net_force_N * tas_array     # P_excess = (T-D)·V [W]
    
    # Fuel efficiency: η = Δm_fuel/Δs [kg/km]
    fuel_efficiency_kg_per_km = np.zeros(n_points)
    for i in range(1, n_points):
        if i < len(distance_km) and distance_km[i] > distance_km[i-1]:
            dist_diff = distance_km[i] - distance_km[i-1]
            fuel_diff = descent_result.cumFuel_kg[i] - descent_result.cumFuel_kg[i-1] if i < len(descent_result.cumFuel_kg) else 0.0
            fuel_efficiency_kg_per_km[i] = fuel_diff / dist_diff if dist_diff > 0 else 0.0
    
    # Get summary statistics and initial/target states
    total_time = descent_result.total_time_s if hasattr(descent_result, 'total_time_s') else time_s[-1] if len(time_s) > 0 else 0.0
    total_fuel = descent_result.total_fuel_consumed_kg if hasattr(descent_result, 'total_fuel_consumed_kg') else descent_result.cumFuel_kg[-1] if len(descent_result.cumFuel_kg) > 0 else 0.0
    final_mass = descent_result.final_mass_kg if hasattr(descent_result, 'final_mass_kg') else weight_array[-1] if len(weight_array) > 0 else 0.0  # Renamed for physics accuracy
    avg_descent_rate = descent_result.average_descent_rate_mps if hasattr(descent_result, 'average_descent_rate_mps') else 0.0
    avg_fuel_flow = descent_result.average_fuel_flow_kgps if hasattr(descent_result, 'average_fuel_flow_kgps') else 0.0
    strategy_name = descent_result.strategy_name if hasattr(descent_result, 'strategy_name') else 'Unknown'
    initial_altitude = descent_result.initial_altitude_m if hasattr(descent_result, 'initial_altitude_m') else descent_result.alt_m[0] if len(descent_result.alt_m) > 0 else 0.0
    initial_mach = descent_result.initial_mach if hasattr(descent_result, 'initial_mach') else descent_result.mach[0] if len(descent_result.mach) > 0 else 0.0
    initial_mass = descent_result.initial_mass_kg if hasattr(descent_result, 'initial_mass_kg') else weight_array[0] if len(weight_array) > 0 else 0.0  # Renamed for physics accuracy
    target_altitude = descent_result.target_altitude_m if hasattr(descent_result, 'target_altitude_m') else descent_result.alt_m[-1] if len(descent_result.alt_m) > 0 else 0.0
    target_mach = descent_result.target_mach if hasattr(descent_result, 'target_mach') else descent_result.mach[-1] if len(descent_result.mach) > 0 else 0.0
    
    # Get summary dictionary for additional fields (before use in data dict)
    descent_summary = descent_result.get_summary_dict() if hasattr(descent_result, 'get_summary_dict') else {}
    
    # Calculate additional metrics for consistency
    thrust_per_engine_N = thrust_array / 2.0  # 2 engines assumed
    
    # Distance increment (non-cumulative distance per segment)
    distance_increment_km = np.concatenate([[0.0], np.diff(distance_km)])
    
    # Calculate target distance (0 for descent as it's altitude-driven, not distance-driven)
    target_distance_km = 0.0
    
    # Thrust limited flag: Check if lever is at maximum (≈ 1.0)
    # Using same tolerance as climb module
    lever_array = np.asarray(descent_result.lever, float)
    thrust_limited = np.isclose(lever_array, 1.0, atol=0.01).astype(int)
    
    # Create DataFrame with all available data from descent_result
    # Organized into logical groups: Phase ID -> Time -> Distance -> Position -> Mass -> Forces -> Fuel -> Atmospheric -> Performance -> Engine -> Efficiency -> Mission Parameters
    data = {
        # Phase identification
        'Phase': ['Descent'] * n_points,
        'Strategy Name': [strategy_name] * n_points,
        
        # Time data
        'Cumulative Time (s)': time_s,
        'Time Step (s)': np.asarray(descent_result.dt_s, float),
        'Total Time (s)': np.full(n_points, total_time),
        
        # Distance data
        'Cumulative Distance (km)': distance_km,
        'Distance (km)': distance_increment_km,
        
        # Position/State
        'Altitude (m)': np.asarray(descent_result.alt_m, float),
        'Mach Number': np.asarray(descent_result.mach, float),
        'True Airspeed (m/s)': tas_array,
        'Speed of Sound (m/s)': speed_of_sound_mps,
        
        # Mass
        'Mass (kg)': weight_array,
        
        # Forces
        'Thrust Total (N)': thrust_array,
        'Thrust Per Engine (N)': thrust_per_engine_N,
        'Average Thrust (N)': np.full(n_points, float(np.mean(thrust_array))) if len(thrust_array) > 0 else np.zeros(n_points),
        'Drag (N)': drag_array,
        'Lever Position': np.asarray(descent_result.lever, float),
        
        # Fuel data
        'Fuel Flow (kg/s)': np.asarray(descent_result.mdot_kgps, float),
        'Fuel Flow (kg/h)': np.asarray(descent_result.mdot_kgps, float) * 3600.0,
        'Mass Flow Rate (kg/s)': np.asarray(descent_result.mdot_kgps, float),  # Same as fuel flow
        'Average Fuel Flow (kg/s)': np.full(n_points, avg_fuel_flow),
        'Cumulative Fuel (kg)': np.asarray(descent_result.cumFuel_kg, float),
        'Total Fuel Consumed (kg)': np.full(n_points, total_fuel),
        'Fuel Increment (kg)': np.asarray(descent_result.dFuel_kg, float),
        'Fuel Estimate (kg)': np.full(n_points, total_fuel),  # Total fuel consumed for descent phase
        
        # Atmospheric properties
        'Temperature (K)': np.asarray(descent_result.temperature_K, float),
        'Pressure (Pa)': pressure_Pa,
        'Density (kg/m³)': density_array,
        
        # Performance metrics
        'Specific Excess Power (m/s)': np.asarray(descent_result.Ps_mps, float),
        'Climb Rate (m/s)': climb_rate_mps,
        'Descent Rate (m/s)': np.asarray(descent_result.Ps_mps, float),
        'Average Descent Rate (m/s)': np.full(n_points, avg_descent_rate),
        'Average Descent Rate (m/min)': np.full(n_points, descent_summary.get('avg_descent_rate_mpm', avg_descent_rate * 60.0)),
        'J Cost (kg/m)': np.asarray(descent_result.J_kg_per_m, float) if hasattr(descent_result, 'J_kg_per_m') else np.zeros(n_points),  # Fuel cost density
        'Thrust Limited': thrust_limited,  # Calculated from lever position
        
        # Aerodynamic coefficients
        'Dynamic Pressure (Pa)': dynamic_pressure_Pa,
        'Lift Coefficient': lift_coefficient,
        'Drag Coefficient': drag_coefficient,
        'L/D Ratio': ld_ratio,
        
        # Engine performance
        'TSFC (kg/(N·s))': tsfc_kg_per_N_s,
        'TSFC (g/(kN·s))': tsfc_kg_per_N_s * 1e6,  # Convert to g/(kN·s)
        
        # Force and power metrics
        'Net Force (N)': net_force_N,
        'Thrust Power (W)': thrust_power_W,
        'Drag Power (W)': drag_power_W,
        'Excess Power (W)': excess_power_W,
        'Thrust Power (kW)': thrust_power_W / 1000.0,
        'Drag Power (kW)': drag_power_W / 1000.0,
        
        # Efficiency metrics
        'Fuel Efficiency (kg/km)': fuel_efficiency_kg_per_km,
        
        # Mission parameters
        'Initial Altitude (m)': np.full(n_points, initial_altitude),
        'Initial Mach': np.full(n_points, initial_mach),
        'Initial Mass (kg)': np.full(n_points, initial_mass),
        'Target Altitude (m)': np.full(n_points, target_altitude),
        'Target Mach': np.full(n_points, target_mach),
        'Target Distance (km)': np.full(n_points, target_distance_km),
        'Final Mass (kg)': np.full(n_points, final_mass),
    }
    
    return pd.DataFrame(data)


# ========================================================================
# SECTION 6: CG ANALYSIS DATA EXTRACTION
# ========================================================================

def _prepare_cg_data() -> Optional[pd.DataFrame]:
    """
    Extract CG evolution data from fuel system history.
    
    Data extracted:
        - x_CG(t): Longitudinal CG position [m]
        - m_fuel(t): Total fuel remaining [kg]
        - m_i(t): Individual tank fuel levels [kg] for i=0..4
        - Tank groups: Outer {1,3}, Inner {0,2}, Center {4}
        - Tank CG positions: x_i [m] (constant reference)
    
    Source: FuelSystem singleton history tracker.
    
    Returns:
        DataFrame: CG time series, or None if no history available
    """
    try:
        # Get fuel system instance (uses the existing singleton)
        fuel_system = _get_fuel_system()
        
        # Check if history is available
        if not hasattr(fuel_system, 'history_tracker') or len(fuel_system.history_tracker.cg_history) == 0:
            print("[CG_EXPORT] No CG history available - skipping CG Analysis sheet")
            return None
        
        # Get history data
        cg_history = fuel_system.history_tracker.cg_history
        fuel_consumed_history = fuel_system.history_tracker.fuel_consumed_history
        fuel_remaining_history = fuel_system.history_tracker.fuel_remaining_history
        weight_history = fuel_system.history_tracker.weight_history
        tank_fuel_history = fuel_system.history_tracker.tank_fuel_history
        
        n_points = len(cg_history)
        
        if n_points == 0:
            print("[CG_EXPORT] Empty CG history - skipping CG Analysis sheet")
            return None
        
        # Extract tank fuel levels for each point
        tank_0_history = np.array(tank_fuel_history.get(0, [0.0] * n_points))
        tank_1_history = np.array(tank_fuel_history.get(1, [0.0] * n_points))
        tank_2_history = np.array(tank_fuel_history.get(2, [0.0] * n_points))
        tank_3_history = np.array(tank_fuel_history.get(3, [0.0] * n_points))
        tank_4_history = np.array(tank_fuel_history.get(4, [0.0] * n_points))
        
        # Calculate tank group totals
        outer_tanks_total = tank_1_history + tank_3_history
        inner_tanks_total = tank_0_history + tank_2_history
        center_tank_total = tank_4_history
        
        # Create data dictionary with comprehensive CG analysis
        data = {
            # Index for reference
            'Point Index': np.arange(1, n_points + 1),
            
            # CG Position
            'CG_X Position (m)': np.array(cg_history),
            
            # Fuel consumption summary
            'Fuel Remaining (kg)': np.array(fuel_remaining_history),
            'Fuel Consumed (kg)': np.array(fuel_consumed_history),
            
            # Aircraft mass (if available)
            'Aircraft Mass (kg)': np.array(weight_history) if len(weight_history) > 0 else np.full(n_points, np.nan),
            
            # Individual tank fuel levels
            f'{TANK_NAMES[0]} Fuel (kg)': tank_0_history,
            f'{TANK_NAMES[1]} Fuel (kg)': tank_1_history,
            f'{TANK_NAMES[2]} Fuel (kg)': tank_2_history,
            f'{TANK_NAMES[3]} Fuel (kg)': tank_3_history,
            f'{TANK_NAMES[4]} Fuel (kg)': tank_4_history,
            
            # Tank group totals
            'Outer Tanks Total (kg)': outer_tanks_total,
            'Inner Tanks Total (kg)': inner_tanks_total,
            'Center Tank Total (kg)': center_tank_total,
            
            # Tank CG positions (constant for reference)
            f'{TANK_NAMES[0]} CG (m)': np.full(n_points, TANK_CG_POSITIONS[0]),
            f'{TANK_NAMES[1]} CG (m)': np.full(n_points, TANK_CG_POSITIONS[1]),
            f'{TANK_NAMES[2]} CG (m)': np.full(n_points, TANK_CG_POSITIONS[2]),
            f'{TANK_NAMES[3]} CG (m)': np.full(n_points, TANK_CG_POSITIONS[3]),
            f'{TANK_NAMES[4]} CG (m)': np.full(n_points, TANK_CG_POSITIONS[4]),
            
            # Fuel consumption scenario (constant)
            'Consumption Scenario': [fuel_system.scenario] * n_points,
        }
        
        print(f"[CG_EXPORT] Prepared CG analysis data: {n_points} history points")
        print(f"[CG_EXPORT]   Consumption scenario: {fuel_system.scenario}")
        print(f"[CG_EXPORT]   CG_X range: {min(cg_history):.3f} m to {max(cg_history):.3f} m")
        
        return pd.DataFrame(data)
        
    except Exception as e:
        print(f"[CG_EXPORT] Failed to prepare CG data: {str(e)}")
        import traceback
        traceback.print_exc()
        return None

